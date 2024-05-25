import openpyxl


class BaseUserManagement:
    def __init__(self, workbook_name='chat.xlsx'):
        self.workbook_name = workbook_name
        try:
            self.workbook = openpyxl.load_workbook(self.workbook_name)
            self.users_sheet = self.workbook['users']
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.users_sheet = self.workbook.active
            self.users_sheet.title = 'users'
            self.users_sheet.append(['ID', 'Username', 'Password'])
            self.workbook.save(self.workbook_name)

    def add_user(self, username, password):
        user_id = self.users_sheet.max_row
        self.users_sheet.append([user_id, username, password])
        self.workbook.save(self.workbook_name)
        print(f"User '{username}' added successfully.")

    def find_user(self, username):
        for row in self.users_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == username:
                return row[0]
        return None


class AdminUserManagement(BaseUserManagement):
    def remove_user(self, username):
        row_index = None
        for idx, row in enumerate(self.users_sheet.iter_rows(min_row=2), start=2):
            if row[1].value == username:
                row_index = idx
                break
        if row_index is not None:
            self.users_sheet.delete_rows(row_index)
            self.workbook.save(self.workbook_name)
            print(f"User '{username}' removed successfully.")
        else:
            print(f"User '{username}' not found.")

    def find_user(self, username):
        user_id = super().find_user(username)
        if user_id is not None:
            print(f"User '{username}' found with ID: {user_id}.")
        else:
            print(f"User '{username}' not found.")


if __name__ == "__main__":
    admin = AdminUserManagement()
    while True:
        admin_choice = input("Do you want to add, remove, find a user, or exit? (add/remove/find/exit): ").strip().lower()

        if admin_choice == "add":
            username = input("Enter new username: ").strip()
            password = input("Enter new password: ").strip()
            admin.add_user(username, password)
        elif admin_choice == "remove":
            username = input("Enter username to remove: ").strip()
            admin.remove_user(username)
        elif admin_choice == "find":
            username = input("Enter username to find: ").strip()
            admin.find_user(username)
        elif admin_choice == "exit":
            break
        else:
            print("Invalid choice. Please enter 'add', 'remove', 'find', or 'exit'.")
