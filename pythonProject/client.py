import socket
import threading
import openpyxl

HOST = '192.168.0.113'
PORT = 12346

try:
    workbook = openpyxl.load_workbook('chat.xlsx')
    users_sheet = workbook['users']
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    users_sheet = workbook.active
    users_sheet.title = 'users'
    users_sheet.append(['ID', 'Username', 'Password'])
    workbook.save('chat.xlsx')


def find_user(username, password):
    for row in users_sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == username and row[2] == password:
            return row[0]
    return None


def register_user(username, password):
    user_id = users_sheet.max_row
    users_sheet.append([user_id, username, password])
    workbook.save('chat.xlsx')
    return user_id


def receive_messages(client_socket, stop_event):
    while not stop_event.is_set():
        try:
            message = client_socket.recv(1024).decode('utf-8')
            if message:
                print(message)
            else:
                break
        except Exception as e:
            if not stop_event.is_set():
                print(f"Error receiving message: {e}")
            break


# Define a function to handle user authentication
def handle_authentication():
    while True:
        choice = input("Do you want to log in or register? (login/register): ").strip().lower()
        if choice in ["login", "register"]:
            break
        print("Invalid choice. Please enter 'login' or 'register'.")

    while True:
        username = input("Enter your username: ").strip()
        password = input("Enter your password: ").strip()

        if choice == "register":
            user_id = register_user(username, password)
            print(f"Registration successful. Your user ID is: {user_id}")
            return username, password, user_id
        else:
            user_id = find_user(username, password)
            if user_id is not None:
                return username, password, user_id
            else:
                print("Invalid username or password")
                retry_choice = input("Do you want to try again or exit? (try again/exit): ").strip().lower()
                if retry_choice == "exit":
                    return None, None, None
                elif retry_choice == "try again":
                    return handle_authentication()


def start_client():
    username, password, user_id = handle_authentication()
    if user_id is None:
        return

    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        client.connect((HOST, PORT))
        client.send(username.encode('utf-8'))
        print('Connected to the server.')

        stop_event = threading.Event()
        receive_thread = threading.Thread(target=receive_messages, args=(client, stop_event))
        receive_thread.start()

        while True:
            message = input()
            if message == "@exit":
                client.send(f"{username} has left the chat.".encode('utf-8'))
                stop_event.set()
                break
            elif message.startswith('@'):
                recipient, msg = message.split(' ', 1)
                client.send(f"{recipient} {msg}".encode('utf-8'))
            else:
                client.send(message.encode('utf-8'))
    except Exception as e:
        print(f"Failed to connect to the server: {e}")
    finally:
        client.close()
        stop_event.set()
        receive_thread.join()


if __name__ == "__main__":
    start_client()
