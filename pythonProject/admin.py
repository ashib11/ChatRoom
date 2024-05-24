import openpyxl

try:
    workbook = openpyxl.load_workbook('chat.xlsx')
    users_sheet = workbook['users']
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    users_sheet = workbook.active
    users_sheet.title = 'users'
    users_sheet.append(['ID', 'Username', 'Password'])
    workbook.save('chat.xlsx')


def add_user(username, password):
    user_id = users_sheet.max_row
    users_sheet.append([user_id, username, password])
    workbook.save('chat.xlsx')
    print(f"User '{username}' added successfully.")


def remove_user(username):
    row_index = None
    for idx, row in enumerate(users_sheet.iter_rows(min_row=2), start=2):
        if row[1].value == username:
            row_index = idx
            break

    if row_index is not None:
        users_sheet.delete_rows(row_index)
        workbook.save('chat.xlsx')
        print(f"User '{username}' removed successfully.")
    else:
        print(f"User '{username}' not found.")


if __name__ == "__main__":
    while True:
        admin_choice = input("Do you want to add or remove a user? (add/remove/exit): ").strip().lower()

        if admin_choice == "add":
            username = input("Enter new username: ").strip()
            password = input("Enter new password: ").strip()
            add_user(username, password)
        elif admin_choice == "remove":
            username = input("Enter username to remove: ").strip()
            remove_user(username)
        elif admin_choice == "exit":
            break
        else:
            print("Invalid choice. Please enter 'add', 'remove', or 'exit'.")
