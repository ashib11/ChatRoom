import socket
import threading
import openpyxl

HOST = '192.168.240.60'
PORT = 12346


class UserManagement:
    def __init__(self, workbook_name='chat.xlsx'):
        self.workbook_name = workbook_name
        try:
            self.workbook = openpyxl.load_workbook(self.workbook_name)
            self.users_sheet = self.workbook['users']
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.users_sheet = self.workbook.active
            self.users_sheet.title = 'users'
            self.users_sheet.append(['ID', 'Username', 'Password'])
            self.workbook.save(self.workbook_name)

    def add_user(self, username, password):
        user_id = self.users_sheet.max_row
        self.users_sheet.append([user_id, username, password])
        self.workbook.save(self.workbook_name)
        return user_id

    def find_user(self, username, password):
        for row in self.users_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == username and row[2] == password:
                return row[0]
        return None


def receive_messages(client_socket, stop_event):
    while not stop_event.is_set():
        try:
            message = client_socket.recv(1024).decode('utf-8')
            if message:
                print(message)
            else:
                break
        except Exception as e:
            if not stop_event.is_set():
                print(f"Error receiving message: {e}")
            break


def handle_authentication(user_mgmt):
    while True:
        choice = input("Do you want to log in or register? (login/register): ").strip().lower()
        if choice in ["login", "register"]:
            break
        print("Invalid choice. Please enter 'login' or 'register'.")

    while True:
        username = input("Enter your username: ").strip()
        password = input("Enter your password: ").strip()

        if choice == "register":
            user_id = user_mgmt.add_user(username, password)
            print(f"Registration successful. Your user ID is: {user_id}")
            return username, password, user_id
        else:
            user_id = user_mgmt.find_user(username, password)
            if user_id is not None:
                return username, password, user_id
            else:
                print("Invalid username or password")
                retry_choice = input("Do you want to try again or exit? (try again/exit): ").strip().lower()
                if retry_choice == "exit":
                    return None, None, None
                elif retry_choice == "try again":
                    handle_authentication(user_mgmt)


def start_client():
    user_mgmt = UserManagement()
    username, password, user_id = handle_authentication(user_mgmt)
    if user_id is None:
        return

    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        client.connect((HOST, PORT))
        client.send(username.encode('utf-8'))
        print('Connected to the server.')

        stop_event = threading.Event()
        receive_thread = threading.Thread(target=receive_messages, args=(client, stop_event))
        receive_thread.start()

        while True:
            message = input()
            if message == "@exit":
                client.send(f"{username} has left the chat.".encode('utf-8'))
                stop_event.set()
                break
            elif message.startswith('@'):
                recipient, msg = message.split(' ', 1)
                client.send(f"{recipient} {msg}".encode('utf-8'))
            else:
                client.send(message.encode('utf-8'))
    except Exception as e:
        print(f"Failed to connect to the server: {e}")
    finally:
        client.close()
        stop_event.set()
        receive_thread.join()


if __name__ == "__main__":
    start_client()
